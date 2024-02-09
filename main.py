from openpyxl import load_workbook
from datetime import timedelta
import re
import tkinter as tk
from tkinter import simpledialog
from tkinter import filedialog
import openpyxl

buttons = []  # Список для хранения ссылок на кнопки


def sum_time(time_list):
    total_time = timedelta()

    for time_str in time_list:
        if re.match(r'^\d+:\d{2}$', time_str):
            minutes, seconds = map(int, time_str.split(':'))
            total_time += timedelta(minutes=minutes, seconds=seconds)

    return total_time


def convert_to_minutes_seconds(time_delta):
    # Получаем общее количество секунд
    total_seconds = time_delta.total_seconds()

    # Вычисляем минуты и секунды
    total_minutes = total_seconds // 60
    total_seconds = total_seconds % 60

    return int(total_minutes), int(total_seconds)


def browse_file():
    for button in buttons:
        button.destroy()  # Уничтожаем каждую кнопку из списка
    buttons.clear()  # Очищаем список ссылок на кнопки

    file_path = filedialog.askopenfilename()
    entry0.delete(0, tk.END)  # Очищаем поле ввода
    entry0.insert(0, file_path)  # Вставляем выбранный путь в поле ввода

    wb = openpyxl.load_workbook(file_path)
    sheet_names = wb.sheetnames

    selected_workbook = tk.StringVar(root)

    frame1 = tk.Frame(root)
    frame1.pack(fill=tk.X)
    buttons.append(frame1)

    label1 = tk.Label(frame1, text="Выберите лист из выпадающего меню:")
    label1.pack(side=tk.LEFT, padx=(10, 0))

    workbook_menu = tk.OptionMenu(frame1, selected_workbook, *sheet_names,
                                  command=lambda selected: on_workbook_menu_select(file_path, selected))
    workbook_menu.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)


def on_workbook_menu_select(file_path, selected_workbook):
    frame2 = tk.Frame(root)
    frame2.pack(fill=tk.X, pady=5)
    buttons.append(frame2)

    label2 = tk.Label(frame2, text=f"Введите номер столбца в листе {selected_workbook}:")
    label2.pack(side=tk.LEFT, padx=(10, 0))

    entry1 = tk.Entry(frame2, width=5)
    entry1.pack(side=tk.LEFT, padx=10)

    confirm_button = tk.Button(frame2, text="Подсчитать",
                               command=lambda: confirm(file_path, selected_workbook, entry1, result1, result2))
    confirm_button.pack(side=tk.LEFT)

    frame_result1 = tk.Frame(root)
    frame_result1.pack(fill=tk.BOTH, pady=5)
    buttons.append(frame_result1)

    result_label1 = tk.Label(frame_result1, text=f"Общее время:")
    result_label1.pack(side=tk.LEFT, padx=(10, 0))
    result1 = tk.Entry(frame_result1, width=10)
    result1.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)

    frame_result2 = tk.Frame(root)
    frame_result2.pack(fill=tk.BOTH, pady=5)
    buttons.append(frame_result2)

    result_label2 = tk.Label(frame_result2, text=f"минуты:секунды")
    result_label2.pack(side=tk.LEFT, padx=(10, 0))
    result2 = tk.Entry(frame_result2, width=10)
    result2.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)


def confirm(file_path, sheet_name, entry1, result1, result2):
    column = entry1.get()
    result1.delete(0, tk.END)  # Очищаем поле result1
    result2.delete(0, tk.END)  # Очищаем поле result2
    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook[sheet_name]

        time_values = []

        for row in sheet.iter_rows(values_only=True):
            time_values.append(row[int(column) - 1])  # Вычитаем 1, так как индексы в Python начинаются с 0

        total_time = sum_time(time_values)
        result1.insert(0, total_time)  # Вставляем в поле result1

        total_minutes, total_seconds = convert_to_minutes_seconds(total_time)
        result2.insert(0, f'{total_minutes}:{total_seconds}')  # Вставляем в поле result2

    except Exception as e:
        print(f"Произошла ошибка при чтении файла: {e}")
        return []


if __name__ == "__main__":
    root = tk.Tk()
    root.title("Выбор файла")

    root.geometry("400x300")

    label0 = tk.Label(root, text="Выберите файл")
    label0.pack(pady=5)

    frame0 = tk.Frame(root)
    frame0.pack(fill=tk.X)

    browse_button = tk.Button(frame0, text="Обзор", command=browse_file)
    browse_button.pack(side=tk.LEFT, padx=(10, 0))

    entry0 = tk.Entry(frame0, width=50)
    entry0.pack(side=tk.LEFT, fill=tk.X, padx=10, expand=True)

    root.mainloop()
